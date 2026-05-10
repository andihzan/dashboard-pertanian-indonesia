"""Parser untuk OJK Statistik Perbankan Indonesia (SPI) Excel.

Source: OJK SPI Desember tiap tahun (2014-2024).
Sheet target: "Kredit LU per Lok.Dati I_X.X.a." — kredit menurut lapangan
usaha per provinsi/Daerah Tingkat I.

Format sheet (2015-2024, .xlsx):
  Row 1-3: Title / kosong
  Row 4: Header kolom (Pertanian, Perikanan, Pertambangan, ...)
  Row 5: Nomor kolom (1, 2, 3, ...)
  Row 6+: Province rows alternating with NPL/NPF rows
    Row N (provinsi):  [- | <Provinsi> | <Pertanian> | <Perikanan> | ...]
    Row N+1 (NPL):     [- | NPL/NPF | <NPL Pertanian> | ... ]

Format sheet (2014, .xls — sheet "Kredit LU per Lok.Dati I_4.8"):
  Row 3: Header kolom (Keterangan | Pertanian | Perikanan | ...)
  Row 5+: Province rows alternating with NPL rows
    Row N (provinsi):  [<Provinsi> | <Pertanian> | <Perikanan> | ...]
    Row N+1 (NPL):     [   NPL | <NPL Pertanian> | ... ]

Output: kredit pertanian (Pertanian + Perikanan) per provinsi per tahun
in Rp Miliar.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from src.provinces import PROVINCE_NAMES

PROVINCE_NORMALIZE = {
    "D.I Yogyakarta": "DI Yogyakarta",
    "D.I. Yogyakarta": "DI Yogyakarta",
    "Daerah Istimewa Yogyakarta": "DI Yogyakarta",
    "DI Yogyakarta ": "DI Yogyakarta",
    "Dki Jakarta": "DKI Jakarta",
    "DKI Jakarta ": "DKI Jakarta",
    "Bangka Belitung": "Kepulauan Bangka Belitung",
    "Kepulauan Bangka Belitung ": "Kepulauan Bangka Belitung",
}


def _normalize_province(s: str) -> str | None:
    if not isinstance(s, str):
        return None
    s = s.strip().rstrip(":")
    if s in PROVINCE_NORMALIZE:
        return PROVINCE_NORMALIZE[s]
    if s in PROVINCE_NAMES:
        return s
    # Try case-insensitive
    s_lower = s.lower()
    for p in PROVINCE_NAMES:
        if p.lower() == s_lower:
            return p
    for k, v in PROVINCE_NORMALIZE.items():
        if k.lower() == s_lower:
            return v
    return None


def _find_kredit_lu_sheet_names(wb_sheetnames: list[str]) -> str | None:
    """Cari sheet Kredit LU per Lok.Dati I (number prefix bisa berbeda)."""
    for s in wb_sheetnames:
        if s.startswith("Kredit LU per Lok.Dati I") and "BPR" not in s:
            return s
    return None


def _find_kredit_lu_sheet(wb) -> str | None:
    """Wrapper untuk openpyxl workbook."""
    return _find_kredit_lu_sheet_names(wb.sheetnames)


def _extract_rows_from_sheet(rows: list, year: int) -> pd.DataFrame:
    """Shared extraction logic untuk baris-baris sheet (list of tuples)."""
    # Find header row containing "Pertanian"
    pertanian_col = None
    perikanan_col = None
    for ri, row in enumerate(rows[:10]):
        for ci, cell in enumerate(row):
            if isinstance(cell, str):
                cell_lower = cell.lower().strip()
                if pertanian_col is None and "pertanian" in cell_lower and "perburuan" in cell_lower:
                    pertanian_col = ci
                elif perikanan_col is None and cell_lower.startswith("perikanan"):
                    perikanan_col = ci

    if pertanian_col is None:
        return pd.DataFrame()

    out_rows = []
    seen_provs = set()
    for row in rows:
        if not row or len(row) < 3:
            continue
        # Province cell: try col 1 first (2015+), then col 0 (2014)
        for prov_col in [1, 0, 2]:
            if prov_col >= len(row):
                continue
            cell = row[prov_col]
            if isinstance(cell, str):
                canonical = _normalize_province(cell)
                if canonical and canonical not in seen_provs:
                    pert_val = row[pertanian_col] if pertanian_col < len(row) else None
                    perik_val = (row[perikanan_col]
                                 if perikanan_col is not None and perikanan_col < len(row)
                                 else None)
                    if isinstance(pert_val, (int, float)):
                        out_rows.append({
                            "tahun": year, "provinsi": canonical,
                            "jenis_kredit": "Pertanian, Perburuan & Kehutanan",
                            "nilai_miliar_rp": float(pert_val),
                            "sumber": "OJK SPI", "status": "real",
                        })
                    if isinstance(perik_val, (int, float)):
                        out_rows.append({
                            "tahun": year, "provinsi": canonical,
                            "jenis_kredit": "Perikanan",
                            "nilai_miliar_rp": float(perik_val),
                            "sumber": "OJK SPI", "status": "real",
                        })
                    if isinstance(pert_val, (int, float)) or isinstance(perik_val, (int, float)):
                        total = (float(pert_val) if isinstance(pert_val, (int, float)) else 0) + \
                                (float(perik_val) if isinstance(perik_val, (int, float)) else 0)
                        out_rows.append({
                            "tahun": year, "provinsi": canonical,
                            "jenis_kredit": "Total",
                            "nilai_miliar_rp": total,
                            "sumber": "OJK SPI", "status": "real",
                        })
                    seen_provs.add(canonical)
                    break

    return pd.DataFrame(out_rows)


def _parse_xls(xls_path: Path, year: int) -> pd.DataFrame:
    """Parse legacy .xls file using xlrd (OJK SPI 2014 format)."""
    try:
        import xlrd
    except ImportError:
        raise ImportError("xlrd diperlukan untuk file .xls lama: pip install xlrd")

    wb = xlrd.open_workbook(str(xls_path))
    sheet_name = _find_kredit_lu_sheet_names(wb.sheet_names())
    if not sheet_name:
        return pd.DataFrame()

    ws = wb.sheet_by_name(sheet_name)
    # Convert to list-of-tuples, replacing xlrd empty with None
    rows = []
    for ri in range(ws.nrows):
        row = []
        for ci in range(ws.ncols):
            cell = ws.cell(ri, ci)
            # xlrd cell types: 0=empty,1=text,2=number,3=date,4=bool,5=error
            if cell.ctype == 1:
                row.append(str(cell.value))
            elif cell.ctype == 2:
                row.append(cell.value)
            else:
                row.append(None)
        rows.append(tuple(row))

    return _extract_rows_from_sheet(rows, year)


def parse_ojk_spi(xlsx_path: Path, year: int) -> pd.DataFrame:
    """Parse one OJK SPI Excel file (.xlsx atau .xls lama).

    Returns DataFrame with columns:
        tahun, provinsi, jenis_kredit, nilai_miliar_rp, sumber, status
    """
    p = Path(xlsx_path)
    if p.suffix.lower() == ".xls":
        return _parse_xls(p, year)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = _find_kredit_lu_sheet(wb)
    if not sheet_name:
        return pd.DataFrame()
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if len(rows) < 6:
        return pd.DataFrame()

    return _extract_rows_from_sheet(rows, year)


if __name__ == "__main__":
    import sys
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        "data/raw/ojk/spi_des_2024.xlsx"
    )
    year = int(sys.argv[2]) if len(sys.argv) > 2 else 2024
    df = parse_ojk_spi(xlsx, year)
    print(f"\nParsed {len(df)} rows")
    print(df.head(20).to_string(index=False))
    print(f"\nProvinces: {df['provinsi'].nunique()}")
